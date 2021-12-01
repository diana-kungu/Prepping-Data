# Reading named table from Excel

#Imports
from pandas import DataFrame, melt, concat
from openpyxl import load_workbook 
from numpy import where


#Read data
wb = load_workbook(r'.\2021\Week - 48\Input\PD 2021 Wk 48 Input.xlsx', read_only= False)

# Read named tables
ws = wb[wb.sheetnames[0]]

df_lst = []
for k in ws.tables.keys():
    lookup_table = ws.tables[k]
    data = ws[lookup_table.ref]
    rows_list = []

    # Loop through each row and get the values in the cells
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        
        rows_list.append(cols)

    # Create a pandas dataframe from the rows_list.
    # The first row is the column names
    df_lst.append(DataFrame(data=rows_list[1:], index=None, columns=rows_list[0]))
    
cols_names = ['Clean Measure names', 2020,2021, 'Branch']

#Rename df columns
for i, df in enumerate(df_lst):
    df['Branch'] = df.columns[0]
    df.columns = cols_names
    df = df.melt(id_vars =['Clean Measure names', 'Branch'] ,  
                var_name = 'Recorded Year', value_name= 'True Value')
    df_lst[i] = df
 
 
# Merge all stores tables
df = concat(df_lst, ignore_index=True)

# Create True value column for the correct number of zeros for the measures
df['True Value'] = (where(df['Clean Measure names'].str.slice(-3) =='(m)', df['True Value']*1000_000,
                    where(df['Clean Measure names'].str.slice(-3) =='(k)', df['True Value']*1_000,
                        df['True Value']))).astype('int')
df['True Value'] = df['True Value'].map(lambda x : "{:,}".format(x))

# Remove the suffix of the measure (i.e. the (k) or (m) if the measure name has the units)
df['Clean Measure names'] =[x[:-4]  if '(' in x else x for x in df['Clean Measure names']]


#OUTPUT 
df.to_csv(r'.\2021\Week - 48\Output\week_48.csv', index=False)

print('End')

